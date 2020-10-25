# data driven testing using Excel
# the testing driven by data is called data driven testing

# basic operation on excel
# config openpyxl
import openpyxl

path = r'C:\Users\dhruv\Downloads\stocks.xlsx'

workbook = openpyxl.load_workbook(path)

# extract the active sheet from excel file
sheet = workbook.active

# to get sheet dynamically
#sheet1 = workbook.get_sheet_by_name('Sheet1')

# counting rows & cols
rows = sheet.max_row
cols = sheet.max_column

print(rows)
print(cols)

for r in range(1, rows+1):
    for c in range(1, cols+1):

        print(sheet.cell(row=r, column=c).value, '\t \t \t \t',end='')
    print()
